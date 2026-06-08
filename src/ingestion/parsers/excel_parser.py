"""
Excel (.xlsx) 文档解析器。

使用 openpyxl 提取工作表数据。
每个工作表的每行作为一个 LlamaDocument。
"""

import logging
from pathlib import Path
from typing import List

from llama_index.core.schema import Document as LlamaDocument

from src.ingestion.parsers.base_parser import BaseParser, make_doc_id

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Excel 解析器——基于 openpyxl，每行数据一个 LlamaDocument。

    处理所有工作表，metadata 中包含 sheet_name 和 row 位置信息。

    Usage::

        parser = ExcelParser()
        docs = parser.parse("/path/to/data.xlsx")
        for d in docs:
            print(d.metadata["sheet_name"], d.metadata["row_start"], d.text[:100])
    """

    @property
    def supported_format(self) -> str:
        return "xlsx"

    def parse(self, source: str) -> List[LlamaDocument]:
        """解析 .xlsx 文件。

        Args:
            source: .xlsx 文件路径。

        Returns:
            LlamaDocument 列表，每行数据一个文档。
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl 未安装")
            return []

        doc_id = make_doc_id(source)
        doc_name = self._source_name(source)
        file_path = Path(source)

        if not file_path.exists():
            logger.error("Excel 文件不存在: %s", source)
            return []

        try:
            wb = load_workbook(source, read_only=True, data_only=True)
        except Exception as e:
            logger.error("无法打开 Excel 文件 %s: %s", source, e)
            return []

        documents: List[LlamaDocument] = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 收集所有行（含表头）
            all_rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=True):
                # 转换所有单元格为字符串
                str_row = [str(cell) if cell is not None else "" for cell in row]
                # 跳过全空行
                if any(v.strip() for v in str_row):
                    all_rows.append(str_row)

            if len(all_rows) < 2:
                logger.debug("工作表 %s 数据行不足（<2 行），跳过", sheet_name)
                continue

            # 第一行作为列名
            headers = all_rows[0]
            data_rows = all_rows[1:]

            for row_idx, row in enumerate(data_rows):
                parts: List[str] = []
                for col_idx, value in enumerate(row):
                    col_name = (
                        headers[col_idx]
                        if col_idx < len(headers)
                        else f"Column{col_idx}"
                    )
                    parts.append(f"{col_name}: {value}")

                row_text = "\n".join(parts)
                if not row_text.strip():
                    continue

                doc = LlamaDocument(
                    text=row_text,
                    metadata={
                        "doc_id": doc_id,
                        "doc_name": doc_name,
                        "source": source,
                        "format": "xlsx",
                        "sheet_name": sheet_name,
                        "row_start": row_idx + 2,  # 1-based, 跳过表头
                        "row_end": row_idx + 2,
                        "total_rows": len(data_rows),
                        "columns": str(headers),
                    },
                )
                documents.append(doc)
                total_rows += 1

        wb.close()
        logger.info(
            "Excel 解析完成: %s (%d 个工作表, %d 行数据)",
            doc_name, len(wb.sheetnames), total_rows,
        )
        return documents

    # ------------------------------------------------------------------
    # LlamaIndex 回退
    # ------------------------------------------------------------------

    @staticmethod
    def parse_via_llamaindex(source: str) -> List[LlamaDocument]:
        """使用 openpyxl 直接读取（备选，保持接口一致）。

        Args:
            source: 文件路径。

        Returns:
            LlamaDocument 列表。
        """
        # PandasExcelReader 需要 pandas，此处提供纯 openpyxl 回退
        parser = ExcelParser()
        return parser.parse(source)
